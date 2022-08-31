import pandas as pd
from openpyxl import load_workbook

excel = "C:\\Users\\DM\\Desktop\\FAQ.xlsx"
# new_excel = "C:\\Users\\DM\\Desktop\\FAQ_1.xlsx"
new_excel = "C:\\Users\\DM\\Desktop\\805 xlxs Dup\\FAQ_1.xlsx"
data = pd.read_excel(excel)

print(data.columns)
# Index(['Domain', 'Intents', 'Slots', 'NotedText', 'Unnamed: 4', 'Unnamed: 5','Unnamed: 6'],)
print(data)

data = data.drop_duplicates(subset=["NotedText"], keep="first")

data.to_excel(new_excel, index=False, encoding='utf-8')

def duplicate_column(file, newfile, columns):
    data = pd.read_excel(file,sheet_name=0)
    data = data.drop_duplicates(subset=columns, keep="first")
    data.to_excel(newfile, index=False, encoding='utf-8')


wb2 = load_workbook(new_excel)
ws = wb2.create_sheet('mytest',0)
ws.sheet_properties.tabColor = 'ff72BA'
wb2.save('Mytest.xlsx')
wb2.close()

